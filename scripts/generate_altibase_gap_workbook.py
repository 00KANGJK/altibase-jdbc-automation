from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


SRC_PATH = r"C:\Users\TTA\IdeaProjects\altibase-jdbc-automation\gs_a_21_225_testcase.xlsx"
OUT_PATH = r"C:\Users\TTA\IdeaProjects\altibase-jdbc-automation\GS-A-21-225_재분류_및_추가테스트케이스.xlsx"


QUALITY_RULES = [
    {
        "quality": "기능적합성",
        "sub": "기능완전성/정확성",
        "current_count": 781,
        "related": "Schema Object, Non-schema Object, JDBC Driver, Tablespace, Utility, 데이터 타입",
        "status": "중상",
        "gaps": "오류 처리, 경계값, 잘못된 입력, 운영 상태 제약 검증이 상대적으로 부족함",
        "proposal": "정상/오류/경계 시나리오를 1:1로 보강하고 시스템 카탈로그 및 에러코드 검증을 표준화",
        "priority": "상",
    },
    {
        "quality": "신뢰성",
        "sub": "성숙성/결함허용성/복구성",
        "current_count": 93,
        "related": "Backup, Recovery, Transaction, Tablespace(ONLINE/OFFLINE/DISCARD), Replication 일부",
        "status": "중하",
        "gaps": "비정상 종료 후 restart recovery, media recovery, checkpoint/backup 상호배제, 장애 유도 시나리오가 얕음",
        "proposal": "재기동, 장애 복구, 파일 유실, incomplete recovery, replication failover 검증 추가",
        "priority": "상",
    },
    {
        "quality": "성능효율성",
        "sub": "시간효율성/자원효율성/용량성",
        "current_count": 41,
        "related": "Optimizer, JDBC fetch/batch 일부, Tablespace autoextend, MAXROWS",
        "status": "하",
        "gaps": "정량 baseline, 반복 부하, 대량 처리, 메모리/디스크 사용량 계측이 부족함",
        "proposal": "성능 baseline, 대량 INSERT/SELECT, fetch/batch, autoextend, checkpoint/backup 중 자원 사용량 측정 추가",
        "priority": "중상",
    },
    {
        "quality": "보안성",
        "sub": "인증/인가/책임추적성/기밀성",
        "current_count": 133,
        "related": "User, Privilege & Role, Audit, Encrypt",
        "status": "중",
        "gaps": "계정 잠금, 비인가 접근, role 재접속 반영, audit 산출물, syslog 연동 검증이 부족함",
        "proposal": "권한 거부, 계정 정책, audit file/syslog, 암호화 데이터 정합성 검증 추가",
        "priority": "상",
    },
    {
        "quality": "호환성",
        "sub": "상호운용성/공존성",
        "current_count": 270,
        "related": "JDBC Driver, Utility, Communication, Replication",
        "status": "중",
        "gaps": "iSQL/JDBC/유틸 간 결과 일치, 문자셋, 네트워크 재연결, 복제 동기성 검증이 부족함",
        "proposal": "JDBC-isql 결과 일관성, iLoader/aexport round-trip, replication sync/eager-lazy 차이 검증 추가",
        "priority": "중상",
    },
    {
        "quality": "유지보수성",
        "sub": "분석성/시험성",
        "current_count": 0,
        "related": "현 테스트케이스 문서만으로는 직접 측정 어려움",
        "status": "하",
        "gaps": "실패 시 로그 수집, 원인 분석용 표준 출력, 환경 초기화 규약이 문서에 없음",
        "proposal": "자동화 프레임워크에서 SQL, 에러코드, 서버로그, 환경 snapshot 수집을 표준 기능으로 추가",
        "priority": "중",
    },
    {
        "quality": "이식성",
        "sub": "환경적응성",
        "current_count": 5,
        "related": "Communication, Utility, 파일 경로 기반 기능",
        "status": "하",
        "gaps": "RHEL 경로 권한, 설치 경로 변경, 환경변수/locale/timezone 영향 검증이 부족함",
        "proposal": "RHEL 9.6 기준 경로 권한, locale, timezone, alternate path 환경 테스트 추가",
        "priority": "중",
    },
]


ADDITIONAL_CASES = [
    ("Transaction", "Commit", "AUTO_COMMIT OFF", "-", "TC_687_001", "자동 커밋을 끈 세션에서 COMMIT 전까지 다른 세션에 변경사항이 노출되지 않는지 확인",
     "(사전조건) 두 개의 세션 A/B 준비, 테스트 테이블 tx_tb1 생성  (입력) 1. 세션 A에서 AUTOCOMMIT OFF 설정 2. 세션 A에서 INSERT 수행 후 COMMIT 하지 않음 3. 세션 B에서 동일 데이터 조회 4. 세션 A에서 COMMIT 후 세션 B에서 재조회",
     "COMMIT 이전에는 세션 B에서 변경사항이 보이지 않고 COMMIT 이후 조회됨"),
    ("Transaction", "Rollback", "ROLLBACK", "-", "TC_688_001", "ROLLBACK 실행 시 트랜잭션 내 변경사항이 모두 취소되는지 확인",
     "(사전조건) 테스트 테이블 tx_tb2 생성  (입력) 1. AUTOCOMMIT OFF 설정 2. INSERT/UPDATE 수행 3. ROLLBACK 실행 4. 데이터 재조회",
     "ROLLBACK 이전 변경사항이 모두 취소되고 원본 상태로 복원됨"),
    ("Transaction", "Rollback", "SAVEPOINT", "-", "TC_689_001", "SAVEPOINT 지정 후 부분 롤백이 가능한지 확인",
     "(사전조건) 테스트 테이블 tx_tb3 생성  (입력) 1. AUTOCOMMIT OFF 설정 2. 첫 번째 INSERT 수행 3. SAVEPOINT sp1 선언 4. 두 번째 INSERT 수행 5. ROLLBACK TO SAVEPOINT sp1 실행 6. COMMIT 후 조회",
     "SAVEPOINT 이후 작업만 취소되고 SAVEPOINT 이전 작업은 COMMIT 됨"),
    ("Transaction", "Lock", "ROW LOCK", "-", "TC_690_001", "동일 행에 대해 다른 세션이 갱신을 시도할 때 잠금 대기 또는 오류가 발생하는지 확인",
     "(사전조건) 두 개의 세션 A/B 준비, 테스트 테이블 tx_lock1 생성 및 1건 입력  (입력) 1. 세션 A에서 대상 행 UPDATE 후 COMMIT 하지 않음 2. 세션 B에서 동일 행 UPDATE 시도",
     "세션 B는 잠금 해제 전까지 대기하거나 설정된 잠금 정책에 따라 오류가 반환됨"),
    ("Transaction", "Lock", "DEADLOCK", "-", "TC_691_001", "교차 갱신 상황에서 데드락을 감지하고 한 세션을 롤백하는지 확인",
     "(사전조건) 세션 A/B, 테스트 테이블 tx_lock2 생성 및 2건 입력  (입력) 1. 세션 A가 1번 행 UPDATE 2. 세션 B가 2번 행 UPDATE 3. 세션 A가 2번 행 UPDATE 시도 4. 세션 B가 1번 행 UPDATE 시도",
     "데드락이 감지되고 최소 한 세션의 문장이 실패하며 트랜잭션 무결성이 유지됨"),
    ("Transaction", "Commit", "DURABILITY", "-", "TC_692_001", "COMMIT 완료 후 비정상 종료가 발생해도 restart recovery로 데이터가 유지되는지 확인",
     "(사전조건) 테스트 테이블 tx_dur1 생성  (입력) 1. INSERT 후 COMMIT 2. 서버 비정상 종료 유도 3. 서버 재기동 4. 데이터 조회",
     "재기동 시 restart recovery가 수행되고 COMMIT 완료 데이터가 보존됨"),
    ("Transaction", "Checkpoint", "EXPLICIT SAVEPOINT", "-", "TC_693_001", "문장 오류 발생 시 implicit savepoint 에 의해 해당 문장만 롤백되는지 확인",
     "(사전조건) 테스트 테이블 tx_stmt1 생성, UNIQUE 제약 포함  (입력) 1. AUTOCOMMIT OFF 설정 2. 정상 INSERT 수행 3. 중복값 INSERT로 오류 유도 4. 같은 트랜잭션에서 다른 정상 INSERT 수행 5. COMMIT 후 조회",
     "오류가 발생한 문장만 취소되고 같은 트랜잭션 내 다른 정상 문장은 유지됨"),
    ("Transaction", "Commit", "ISOLATION", "-", "TC_694_001", "동시 세션에서 dirty read가 발생하지 않는지 확인",
     "(사전조건) 세션 A/B 준비, 테스트 테이블 tx_iso1 생성  (입력) 1. 세션 A에서 UPDATE 후 COMMIT 하지 않음 2. 세션 B에서 동일 행 조회",
     "세션 B는 커밋되지 않은 변경값을 읽지 못함"),

    ("Backup", "Online Backup", "ARCHIVELOG", "-", "TC_695_001", "ARCHIVELOG 모드에서 데이터베이스 수준 온라인 백업이 성공하는지 확인",
     "(사전조건) ARCHIVELOG 모드, SYSDBA 계정, 백업 디렉터리 준비  (입력) 1. ALTER DATABASE BACKUP DATABASE TO '/backup_dir'; 실행 2. 생성된 백업 파일 확인",
     "온라인 백업이 성공하고 백업 디렉터리에 데이터 파일 및 관련 파일이 생성됨"),
    ("Backup", "Online Backup", "CHECKPOINT MUTUAL EXCLUSION", "-", "TC_696_001", "온라인 백업 수행 중 체크포인트 요청이 동시에 실행되지 않고 직렬화되는지 확인",
     "(사전조건) 온라인 백업 가능한 환경 준비  (입력) 1. 온라인 백업 시작 2. 다른 세션에서 체크포인트 요청 3. 작업 완료 순서 및 로그 확인",
     "온라인 백업과 체크포인트가 동시에 수행되지 않고 순차적으로 처리됨"),
    ("Backup", "Online Backup", "TABLESPACE STATE", "-", "TC_697_001", "DISCARDED 상태의 테이블스페이스는 온라인 백업 대상에서 제외되는지 확인",
     "(사전조건) DISCARDED 상태 테이블스페이스 준비  (입력) 1. 온라인 백업 수행 2. 백업 결과와 로그 확인",
     "DISCARDED 상태 테이블스페이스는 백업되지 않거나 관련 오류/제외 정보가 제공됨"),
    ("Backup", "Offline Backup", "NORMAL SHUTDOWN", "-", "TC_698_001", "정상 종료 후 수행한 오프라인 백업으로 데이터와 로그 파일을 일관되게 복사할 수 있는지 확인",
     "(사전조건) 백업 대상 경로 준비  (입력) 1. 서버 정상 종료 2. 데이터 파일, 로그 파일, 로그 앵커 파일 복사 3. 백업본 파일 목록 확인",
     "오프라인 백업에 필요한 파일들이 일관된 시점으로 복사됨"),
    ("Recovery", "-", "RESTART RECOVERY", "-", "TC_699_001", "비정상 종료 후 STARTUP 시 restart recovery가 자동 수행되는지 확인",
     "(사전조건) 테스트 데이터 생성 후 COMMIT 완료  (입력) 1. 서버 비정상 종료 유도 2. 서버 기동 3. 서버 로그 및 데이터 조회",
     "STARTUP 과정에서 restart recovery가 자동 수행되고 COMMIT 데이터가 복구됨"),
    ("Recovery", "증분 백업 복구", "MEDIA RECOVERY COMPLETE", "-", "TC_700_001", "백업된 디스크 테이블스페이스 데이터 파일 유실 후 complete media recovery가 가능한지 확인",
     "(사전조건) 온라인 또는 오프라인 백업본 준비, 디스크 테이블스페이스 존재  (입력) 1. 데이터 파일 유실 유도 2. 백업본 복원 3. STARTUP CONTROL 4. ALTER DATABASE RECOVER DATABASE 실행 5. STARTUP SERVICE 6. 데이터 조회",
     "complete media recovery가 성공하고 유실 이전 COMMIT 데이터가 복구됨"),
    ("Recovery", "증분 백업 복구", "MEDIA RECOVERY INCOMPLETE", "-", "TC_701_001", "아카이브 로그 일부가 없을 때 incomplete recovery가 가능한지 확인",
     "(사전조건) 특정 시점 백업본과 일부 로그 손실 상황 준비  (입력) 1. STARTUP CONTROL 2. incomplete recovery 절차 수행 3. STARTUP SERVICE 4. 복구 시점 이후 데이터 상태 확인",
     "지정한 시점까지 데이터가 복구되고 이후 시점 변경사항은 반영되지 않음"),
    ("Recovery", "-", "TEMP TABLESPACE FILE LOSS", "-", "TC_702_001", "임시 테이블스페이스 파일 유실 시 복구 절차가 정상 동작하는지 확인",
     "(사전조건) 임시 테이블스페이스 사용 중인 환경 준비  (입력) 1. 임시 파일 유실 유도 2. 매뉴얼 기준 복구 절차 수행 3. 서비스 기동 및 임시 작업 재실행",
     "임시 테이블스페이스가 정상 복구되거나 재생성되고 서비스가 정상 동작함"),
    ("Recovery", "-", "MEMORY CHECKPOINT IMAGE LOSS", "-", "TC_703_001", "메모리 체크포인트 이미지 파일 유실 시 복구 절차가 정상 동작하는지 확인",
     "(사전조건) 메모리 테이블스페이스와 체크포인트 이미지 존재  (입력) 1. 체크포인트 이미지 파일 유실 유도 2. 복구 절차 수행 3. 서비스 기동 및 데이터 확인",
     "매뉴얼에 정의된 절차에 따라 서비스가 기동되고 복구 가능 범위 내 데이터 정합성이 유지됨"),

    ("Non-schema Object", "User", "LOGIN FAILURE", "-", "TC_704_001", "잘못된 비밀번호로 로그인 시도가 거부되는지 확인",
     "(사전조건) 일반 사용자 sec_user1 생성  (입력) 1. 잘못된 비밀번호로 iSQL 또는 JDBC 로그인 시도",
     "인증이 실패하고 세션이 생성되지 않음"),
    ("Non-schema Object", "User", "LOCK/UNLOCK", "-", "TC_705_001", "잠긴 계정은 로그인할 수 없고 잠금 해제 후 다시 로그인할 수 있는지 확인",
     "(사전조건) 테스트 사용자 sec_user2 생성  (입력) 1. 계정 잠금 수행 2. 로그인 시도 3. 계정 잠금 해제 4. 재로그인 시도",
     "잠긴 상태에서는 로그인에 실패하고 잠금 해제 후 로그인에 성공함"),
    ("Non-schema Object", "Privilege & Role", "OBJECT PRIVILEGE DENY", "-", "TC_706_001", "객체 권한이 없는 사용자는 테이블 조회 또는 변경을 수행할 수 없는지 확인",
     "(사전조건) owner_user와 normal_user 생성, owner_user 소유 테이블 생성  (입력) 1. normal_user로 SELECT 또는 INSERT 시도",
     "권한 부족 오류가 발생하고 데이터는 변경되지 않음"),
    ("Non-schema Object", "Privilege & Role", "GRANT/REVOKE IMMEDIACY", "-", "TC_707_001", "권한 회수 후 기존 세션 또는 재접속 세션에서 접근이 차단되는지 확인",
     "(사전조건) 사용자 grant_user에 객체 권한 부여  (입력) 1. grant_user로 접근 성공 확인 2. REVOKE 수행 3. 동일 또는 재접속 세션에서 재시도",
     "권한 회수 이후 접근이 차단됨"),
    ("Non-schema Object", "Privilege & Role", "ROLE RECONNECT", "-", "TC_708_001", "ROLE로 부여된 권한은 재접속 후 적용되는지 확인",
     "(사전조건) role_test 생성, 필요한 권한을 role_test에 부여, 사용자에 role_test 부여  (입력) 1. 권한 부여 직후 기존 세션에서 작업 시도 2. 재접속 후 동일 작업 수행",
     "기존 세션에서는 즉시 반영되지 않을 수 있으며 재접속 후 ROLE 권한이 정상 적용됨"),
    ("Audit", "Statement Audit", "AUDIT FILE", "-", "TC_709_001", "statement audit 수행 시 감사 로그 파일이 생성되는지 확인",
     "(사전조건) audit 관련 속성 설정 가능, AUDIT_LOG_DIR 확인  (입력) 1. 감사 대상 문장에 대해 auditing 설정 2. 대상 SQL 수행 3. 감사 로그 디렉터리 확인",
     "감사 대상 SQL 실행 내역이 감사 로그에 기록됨"),
    ("Audit", "Statement Audit", "AUDIT SYSLOG", "-", "TC_710_001", "Linux syslog 방식으로 감사 로그를 남길 수 있는지 확인",
     "(사전조건) AUDIT_OUTPUT_METHOD를 syslog 방식으로 설정 가능한 환경  (입력) 1. 관련 속성 설정 2. 감사 대상 SQL 수행 3. syslog 또는 rsyslog 로그 확인",
     "감사 로그가 syslog로 기록됨"),
    ("Audit", "Statement Audit", "AUDIT ROTATION", "-", "TC_711_001", "감사 로그 파일 크기 제한 도달 시 후속 로그 파일이 정상 생성되는지 확인",
     "(사전조건) AUDIT_FILE_SIZE를 작은 값으로 설정  (입력) 1. 감사 대상 SQL을 반복 실행 2. 로그 파일 수와 크기 확인",
     "설정한 크기 이후 후속 감사 로그 파일이 생성되고 로그 유실 없이 기록됨"),
    ("Encrypt", "AES", "ROUND-TRIP", "-", "TC_712_001", "AES 암복호화 함수 사용 시 원문이 정확히 복원되는지 확인",
     "(사전조건) 암복호화 테스트 가능한 계정 준비  (입력) 1. 원문 데이터를 AES로 암호화 2. 복호화 3. 원문과 비교",
     "복호화 결과가 원문과 동일함"),

    ("Non-schema Object", "Replication", "LAZY MODE", "-", "TC_713_001", "LAZY 모드에서 원본 서버 커밋 후 대상 서버 반영이 지연될 수 있는지 확인",
     "(사전조건) local/remote 서버, replication pair, 대상 테이블 준비  (입력) 1. LAZY 모드 replication 구성 2. 원본 서버에서 DML COMMIT 3. 원본/대상 반영 시점 비교",
     "원본 서버 커밋은 즉시 완료되며 대상 서버 반영은 지연될 수 있음"),
    ("Non-schema Object", "Replication", "EAGER MODE", "-", "TC_714_001", "EAGER 모드에서 대상 서버 적용 확인 후 원본 트랜잭션이 커밋되는지 확인",
     "(사전조건) eager mode replication 환경 준비  (입력) 1. 원본 서버에서 DML 수행 2. 커밋 완료 시점과 대상 서버 반영 시점 비교",
     "대상 서버 적용 확인 후 원본 트랜잭션이 커밋되어 데이터 일관성이 보장됨"),
    ("Non-schema Object", "Replication", "SYNC", "-", "TC_715_001", "ALTER REPLICATION ... SYNC 수행 시 원본 데이터가 대상 서버와 동기화되는지 확인",
     "(사전조건) 원본/대상 서버 간 동일 구조 테이블 준비, 일부 데이터 차이 존재  (입력) 1. ALTER REPLICATION ... SYNC 수행 2. 양쪽 데이터 비교",
     "동기화 후 대상 서버 데이터가 원본 서버와 일치함"),
    ("Non-schema Object", "Replication", "DDL CLONE", "-", "TC_716_001", "복제 대상 테이블에 DDL 변경 후 동일 이름 컬럼만 복제되는지 확인",
     "(사전조건) 복제 환경 및 대상 테이블 준비  (입력) 1. 원본 또는 양측에 스키마 변경 DDL 수행 2. 동일 이름 컬럼에 대해 DML 수행 3. 대상 데이터 확인",
     "매뉴얼 제약에 따라 동일 이름 컬럼 데이터만 복제되고 비대응 컬럼은 복제 대상에서 제외됨"),
    ("Non-schema Object", "Replication", "CHARSET", "-", "TC_717_001", "복제 서버 간 문자셋이 다를 경우 구성 또는 동기화가 제한되는지 확인",
     "(사전조건) 문자셋이 다른 서버 환경 또는 시뮬레이션 가능 환경 준비  (입력) 1. replication 구성 시도 2. 오류 또는 제한 사항 확인",
     "문자셋 불일치 환경에서는 복제 구성 또는 정상 동작이 제한됨"),
    ("Non-schema Object", "Replication", "FAILOVER CONSISTENCY", "-", "TC_718_001", "복제 환경에서 active 서버 장애 후 standby 서버 데이터 일관성이 유지되는지 확인",
     "(사전조건) active-standby 복제 환경 준비  (입력) 1. active 서버에서 DML COMMIT 2. standby 반영 확인 3. active 장애 유도 4. standby에서 조회",
     "장애 직전까지 반영된 데이터 범위 내에서 standby 서버가 일관된 데이터를 제공함"),
    ("Non-schema Object", "Replication", "SEQUENCE REPLICATION", "-", "TC_719_001", "sequence replication 구성 시 양쪽 서버에서 중복 없는 시퀀스 값이 유지되는지 확인",
     "(사전조건) sequence replication 환경 준비  (입력) 1. 양쪽 서버에서 sequence 사용 2. 생성된 값 비교",
     "시퀀스 값이 중복 없이 유지되고 fail-over 상황에서도 일관성이 확보됨"),
    ("Non-schema Object", "Replication", "OFFLINE OPTION", "-", "TC_720_001", "일시적인 네트워크 단절 이후 미전송 로그가 재적용되는지 확인",
     "(사전조건) replication 환경과 네트워크 제어 가능 환경 준비  (입력) 1. 네트워크 단절 유도 2. 원본 서버에서 DML 수행 3. 네트워크 복구 4. 대상 서버 데이터 확인",
     "단절 동안 누적된 변경사항이 복구 후 재적용되어 데이터가 동기화됨"),

    ("Utility", "Data Loader(iLoader)", "LOGICAL BACKUP/RESTORE", "-", "TC_721_001", "iLoader로 백업한 테이블을 복원하여 데이터가 일치하는지 확인",
     "(사전조건) 대상 테이블과 샘플 데이터 준비  (입력) 1. iLoader export 수행 2. 원본 테이블 삭제 또는 초기화 3. iLoader import 수행 4. 데이터 비교",
     "복원된 테이블의 데이터와 메타정보가 백업 시점과 일치함"),
    ("Utility", "aexport", "DDL/DATA EXPORT", "-", "TC_722_001", "aexport 수행 시 테이블 생성 스크립트와 데이터 파일이 모두 생성되는지 확인",
     "(사전조건) export 대상 테이블 준비  (입력) 1. aexport 실행 2. 생성 산출물 목록 확인",
     "테이블 생성 스크립트, 인덱스 정보, 데이터 파일이 정상 생성됨"),
    ("Utility", "SQL Query Tool(iSQL)", "ERROR CODE", "-", "TC_723_001", "iSQL에서 잘못된 SQL 실행 시 적절한 오류 코드와 메시지를 반환하는지 확인",
     "(사전조건) iSQL 접속 가능  (입력) 1. 문법 오류 SQL 또는 존재하지 않는 객체 조회 실행",
     "실패 원인을 식별할 수 있는 오류 코드와 메시지가 출력됨"),
    ("Communication", "TCP/IP", "RECONNECT", "-", "TC_724_001", "일시적인 네트워크 단절 후 클라이언트가 재접속하여 정상 질의를 수행할 수 있는지 확인",
     "(사전조건) 서버-클라이언트 간 네트워크 제어 가능  (입력) 1. 정상 질의 수행 2. 네트워크 단절 3. 연결 실패 확인 4. 네트워크 복구 후 재접속 5. 질의 재실행",
     "단절 중에는 연결이 실패하고 복구 후 재접속하여 질의를 정상 수행함"),
    ("Communication", "TCP/IP", "CONNECTION LIMIT", "-", "TC_725_001", "동시 접속 수가 증가해도 허용 범위 내에서 연결과 질의가 안정적으로 수행되는지 확인",
     "(사전조건) 부하 생성 가능한 클라이언트 준비  (입력) 1. 다수 세션 동시 접속 2. 간단한 SELECT 반복 수행 3. 오류 및 응답 확인",
     "허용 범위 내 동시 접속에서 연결 실패나 비정상 종료 없이 질의가 수행됨"),

    ("Optimizer", "Explain Plan", "PLAN CONSISTENCY", "-", "TC_726_001", "EXPLAIN PLAN 결과가 인덱스 생성 전후에 기대한 방식으로 변경되는지 확인",
     "(사전조건) 충분한 데이터가 있는 테스트 테이블 준비  (입력) 1. 인덱스 없는 상태에서 EXPLAIN PLAN 수행 2. 인덱스 생성 3. 동일 질의 EXPLAIN PLAN 재수행",
     "인덱스 생성 후 실행 계획이 변경되어 인덱스 활용 여부를 확인할 수 있음"),
    ("Optimizer", "SQL Plan Cache", "INVALIDATION", "-", "TC_727_001", "테이블 구조 변경 후 SQL plan cache가 무효화되거나 재생성되는지 확인",
     "(사전조건) plan cache 관찰 가능한 환경 준비  (입력) 1. 동일 SQL 반복 실행 2. 관련 테이블 구조 변경 3. SQL 재실행 4. plan cache 상태 확인",
     "구조 변경 후 기존 plan cache가 무효화되거나 새 계획으로 재생성됨"),
    ("Schema Object", "(Non-Partitioned) Table", "BOUNDARY VALUE", "-", "TC_728_001", "VARCHAR, NUMBER 등 컬럼에 허용 최대 길이 또는 정밀도 경계값 입력이 가능한지 확인",
     "(사전조건) 경계값 검증용 테스트 테이블 생성  (입력) 1. 허용 최대 길이 또는 정밀도의 값 입력 2. 저장 결과 조회",
     "허용 범위의 경계값은 정상 저장됨"),
    ("Schema Object", "(Non-Partitioned) Table", "OUT-OF-RANGE", "-", "TC_729_001", "컬럼 정의 범위를 초과한 값 입력 시 오류가 발생하는지 확인",
     "(사전조건) 경계값 검증용 테스트 테이블 생성  (입력) 1. 최대 길이 또는 정밀도를 초과한 값 입력 시도",
     "범위를 초과한 값은 거부되고 적절한 오류가 반환됨"),
    ("JDBC Driver", "Connection", "INVALID CREDENTIAL", "-", "TC_730_001", "JDBC Connection이 잘못된 계정 정보로 생성되지 않는지 확인",
     "(사전조건) Altibase JDBC 드라이버와 접속 URL 준비  (입력) 1. 잘못된 사용자명 또는 비밀번호로 getConnection 호출",
     "SQLException이 발생하고 Connection 객체가 생성되지 않음"),
    ("JDBC Driver", "PreparedStatement", "BATCH ROLLBACK", "-", "TC_731_001", "batch 실행 중 제약조건 오류가 발생할 때 트랜잭션 정책에 맞게 롤백 또는 부분 반영되는지 확인",
     "(사전조건) UNIQUE 제약이 있는 테스트 테이블 준비  (입력) 1. PreparedStatement batch에 정상 데이터와 중복 데이터를 함께 추가 2. executeBatch 수행 3. 결과와 데이터 상태 확인",
     "Altibase와 JDBC 트랜잭션 정책에 맞는 배치 결과가 반환되고 데이터 정합성이 유지됨"),
    ("JDBC Driver", "ResultSet", "FETCH SIZE", "-", "TC_732_001", "fetch size 변경 시 대량 조회 결과의 정확성과 자원 사용이 안정적으로 유지되는지 확인",
     "(사전조건) 대량 데이터가 있는 테스트 테이블 준비  (입력) 1. 서로 다른 fetch size로 동일 SELECT 수행 2. 결과 건수와 오류 여부 비교",
     "fetch size 값이 달라도 결과 정확성은 동일하고 비정상 오류 없이 조회됨"),
]


def auto_fit(ws):
    for col_cells in ws.columns:
        length = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            length = max(length, len(value))
        ws.column_dimensions[col_letter].width = min(max(length + 2, 12), 45)


def copy_header_and_layout(src_ws, dst_ws):
    for col_letter, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = dim.width
    for row_idx in range(1, 6):
        dst_ws.row_dimensions[row_idx].height = src_ws.row_dimensions[row_idx].height
        for col_idx in range(1, 12):
            src_cell = src_ws.cell(row=row_idx, column=col_idx)
            dst_cell = dst_ws.cell(row=row_idx, column=col_idx, value=src_cell.value)
            if src_cell.has_style:
                dst_cell.font = copy(src_cell.font)
                dst_cell.fill = copy(src_cell.fill)
                dst_cell.border = copy(src_cell.border)
                dst_cell.alignment = copy(src_cell.alignment)
                dst_cell.number_format = src_cell.number_format
                dst_cell.protection = copy(src_cell.protection)
    for merged in src_ws.merged_cells.ranges:
        merged_str = str(merged)
        if any(merged_str.startswith(prefix) for prefix in ("A1:", "A2:", "A3:", "A4:", "A5:", "B1:", "B2:", "B3:", "B4:", "B5:", "C1:", "C2:", "C3:", "C4:", "C5:", "D1:", "D2:", "D3:", "D4:", "D5:", "E1:", "E2:", "E3:", "E4:", "E5:", "F1:", "F2:", "F3:", "F4:", "F5:", "G1:", "G2:", "G3:", "G4:", "G5:", "H1:", "H2:", "H3:", "H4:", "H5:", "I1:", "I2:", "I3:", "I4:", "I5:", "J1:", "J2:", "J3:", "J4:", "J5:", "K1:", "K2:", "K3:", "K4:", "K5:")):
            dst_ws.merge_cells(merged_str)


def create_quality_sheet(wb):
    ws = wb.create_sheet("품질재분류")
    headers = ["품질특성", "하위특성", "기존 TC 수", "관련 기능군", "현재 판단", "부족 영역", "보강 제안", "우선순위"]
    ws.append(headers)
    for row in QUALITY_RULES:
        ws.append([
            row["quality"],
            row["sub"],
            row["current_count"],
            row["related"],
            row["status"],
            row["gaps"],
            row["proposal"],
            row["priority"],
        ])
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    auto_fit(ws)
    return ws


def create_additional_cases_sheet(wb, src_ws):
    ws = wb.create_sheet("추가 테스트케이스")
    copy_header_and_layout(src_ws, ws)
    ws["A1"] = "GS-A-21-225 추가 테스트케이스"
    for row_idx in (2, 3):
        for col_idx in range(8, 12):
            ws.cell(row=row_idx, column=col_idx, value=None)
    ws["H2"] = "작성자: Codex"
    ws["H3"] = "작성일: 2026.04.14"
    start_row = 6
    prev = ["", "", "", ""]
    for idx, case in enumerate(ADDITIONAL_CASES, start=start_row):
        l1, l2, l3, l4, tc_id, scenario, test_input, expected = case
        values = [l1, l2, l3, l4, tc_id, scenario, test_input, expected, "", "", "추가 작성"]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=idx, column=col_idx, value=value)
            src_style = src_ws.cell(row=6, column=min(col_idx, 11))
            if src_style.has_style:
                cell.font = copy(src_style.font)
                cell.fill = copy(src_style.fill)
                cell.border = copy(src_style.border)
                cell.alignment = copy(src_style.alignment)
                cell.number_format = src_style.number_format
                cell.protection = copy(src_style.protection)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for level in range(4):
            if values[level] == prev[level]:
                ws.cell(row=idx, column=level + 1, value="")
            else:
                prev = values[:4]
                break
    return ws


def main():
    wb = load_workbook(SRC_PATH)
    src_ws = wb.worksheets[0]
    create_quality_sheet(wb)
    create_additional_cases_sheet(wb, src_ws)
    wb.save(OUT_PATH)
    print(OUT_PATH)


if __name__ == "__main__":
    main()
